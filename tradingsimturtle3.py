import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
#always invests half of current cash
canTradeLong = True
canTradeShort = True 
startingCash = 100000
currentCash = 100000
longStockOnHand = 0 
shortStockOnHand = 0 

#workbook stuff
wb = openpyxl.load_workbook('DOW-2turtle-macd.xlsx', data_only=True)

sheet = wb.get_sheet_by_name('turtle py all')

#h=8 j=10 l=12 m=13 n=14 o=15 
print('Starting Cash for 1972-1982:', startingCash)
#1972-1982
for row in range(2, 2675):
		for col in range (10,12):
		if sheet.cell(row=row, column=col).value == "exit long":
			#print(row, col, "sell long")
			x= 0
			x = longStockOnHand*sheet.cell(row=row, column=8).value
			currentCash = currentCash + x
			longStockOnHand = 0
			canTradeLong = True
			#print(longStockOnHand, x, currentCash)	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value == "exit short":
			#print(row, col, "exit short")	
			x = 0
			x = shortStockOnHand*sheet.cell(row=row, column =8).value
			currentCash -= x
			#print("ssoh:",shortStockOnHand, x)
			shortStockOnHand = 0
			canTradeShort = True
			#print(shortStockOnHand, x, currentCash)
		if sheet.cell(row=row, column=col).value =="long" and canTradeLong == True:
			#print(row,col,"buy long")
			longStockOnHand = longStockOnHand + ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash -= (currentCash/2)
			#print(longStockOnHand, currentCash)
			canTradeLong = False	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value =="short" and canTradeShort == True:
			#print(row, col, "sell short")
			shortStockOnHand += ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash += (currentCash/2)		
			#print(shortStockOnHand, currentCash)
			canTradeShort = False 
print("Conclusion 1972-1982: ")
#print("Starting Cash:", startingCash)
print("Current Cash:", currentCash)
if startingCash > currentCash:	
	print("Loss:", startingCash - currentCash)
elif startingCash < currentCash:
	print("Profit:", currentCash - startingCash)
else:
	print("Breakeven")

#1983-1993
print('')
print('Starting Cash for 1983-1993: ', currentCash)
start83 = currentCash
for row in range(2675, 5457):
	for col in range (10,12):
		if sheet.cell(row=row, column=col).value == "exit long":
			#print(row, col, "sell long")
			x= 0
			x = longStockOnHand*sheet.cell(row=row, column=8).value
			currentCash = currentCash + x
			longStockOnHand = 0
			canTradeLong = True
			#print(longStockOnHand, x, currentCash)	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value == "exit short":
			#print(row, col, "exit short")
			x = 0
			x = shortStockOnHand*sheet.cell(row=row, column =8).value
			currentCash -= x
			#print("ssoh:",shortStockOnHand, x)
			shortStockOnHand = 0
			canTradeShort = True
			#print(shortStockOnHand, x, currentCash)
		if sheet.cell(row=row, column=col).value =="long" and canTradeLong == True:
			#print(row,col,"buy long")
			longStockOnHand = longStockOnHand + ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash -= (currentCash/2)
			#print(longStockOnHand, currentCash)
			canTradeLong = False	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value =="short" and canTradeShort == True:
			#print(row, col, "sell short")
			shortStockOnHand += ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash += (currentCash/2)
			#print(shortStockOnHand, currentCash)
			canTradeShort = False 
print("Conclusion 1983-1993:")
#print("Starting Cash:", start83)
print("Current Cash:", currentCash)
if start83 > currentCash:
	print("Loss:", start83 - currentCash)
elif start83 < currentCash:
	print("Profit:", currentCash - start83)
else:
	print("Breakeven")
#1994-2004
print('')
print("Start for 1994-2004:", currentCash)
start94 = currentCash
for row in range(5457, 8228):
	for col in range (10,12):
		if sheet.cell(row=row, column=col).value == "exit long":
			#print(row, col, "sell long")
			x= 0
			x = longStockOnHand*sheet.cell(row=row, column=8).value
			currentCash = currentCash + x
			longStockOnHand = 0
			canTradeLong = True
			#print(longStockOnHand, x, currentCash)	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value == "exit short":
			#print(row, col, "exit short")
			x = 0
			x = shortStockOnHand*sheet.cell(row=row, column =8).value
			currentCash -= x
			#print("ssoh:",shortStockOnHand, x)
			shortStockOnHand = 0
			canTradeShort = True
			#print(shortStockOnHand, x, currentCash)
		if sheet.cell(row=row, column=col).value =="long" and canTradeLong == True:
			#print(row,col,"buy long")
			longStockOnHand = longStockOnHand + ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash -= (currentCash/2)
			#print(longStockOnHand, currentCash)
			canTradeLong = False	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value =="short" and canTradeShort == True:
			#print(row, col, "sell short")
			shortStockOnHand += ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash += (currentCash/2)
			#print(shortStockOnHand, currentCash)
			canTradeShort = False 

print("Conclusion 1994-2004:")
#print("Starting Cash:", start94)
print("Current Cash:", currentCash)
if start94 > currentCash:
	print("Loss:", start94 - currentCash)
elif start94 < currentCash:
	print("Profit:", currentCash - start94)
else:
	print("Breakeven")

#2005-2016
print("")
print("Starting Cash for 2005-2016:", currentCash)
start05 = currentCash
print(canTradeLong)
for row in range(8228, 11249):
	for col in range (10,12):
		if sheet.cell(row=row, column=col).value == "exit long":
			#print(row, col, "sell long")
			x= 0
			x = longStockOnHand*sheet.cell(row=row, column=8).value
			currentCash = currentCash + x
			longStockOnHand = 0
			canTradeLong = True
			#print(longStockOnHand, x, currentCash)	
			#print(canTradeLong)
			#print('el')
		if sheet.cell(row=row, column=col).value == "exit short":
			#print(row, col, "exit short")
			x = 0
			x = shortStockOnHand*sheet.cell(row=row, column =8).value
			currentCash -= x
			#print("ssoh:",shortStockOnHand, x)
			shortStockOnHand = 0
			canTradeShort = True
			#print(shortStockOnHand, x, currentCash)
			#print('es')
		if sheet.cell(row=row, column=col).value =="long" and canTradeLong == True:
			#print(row,col,"buy long")
			longStockOnHand = longStockOnHand + ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash -= (currentCash/2)
			#print(longStockOnHand, currentCash)
			canTradeLong = False	
			#print(canTradeLong)
			#print('l')
		if sheet.cell(row=row, column=col).value =="short" and canTradeShort == True:
			#print(row, col, "sell short")
			shortStockOnHand += ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash += (currentCash/2)
			#print(shortStockOnHand, currentCash)
			canTradeShort = False
			#print('ss')
print("Conclusion 2005-2016:")
#print("Starting Cash:", start05)
print("Current Cash:", currentCash)
if start05 > currentCash:
	print("Loss:", start05 - currentCash)
elif start05 < currentCash:
	print("Profit:", currentCash - start05)
else:
	print("Breakeven") 
#2017
print("")
print("Starting Cash for 2017:",currentCash)
start17 = currentCash
for row in range(11249, sheet.max_row+1):
	for col in range (10,12):
		if sheet.cell(row=row, column=col).value == "exit long":
			#print(row, col, "sell long")
			x= 0
			x = longStockOnHand*sheet.cell(row=row, column=8).value
			currentCash = currentCash + x
			longStockOnHand = 0
			canTradeLong = True
			#print(longStockOnHand, x, currentCash)	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value == "exit short":
			#print(row, col, "exit short")
			x = 0
			x = shortStockOnHand*sheet.cell(row=row, column =8).value
			currentCash -= x
			#print("ssoh:",shortStockOnHand, x)
			shortStockOnHand = 0
			canTradeShort = True
			#print(shortStockOnHand, x, currentCash)
		if sheet.cell(row=row, column=col).value =="long" and canTradeLong == True:
			#print(row,col,"buy long")
			longStockOnHand = longStockOnHand + ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash -= (currentCash/2)
			#print(longStockOnHand, currentCash)
			canTradeLong = False	
			#print(canTradeLong)
		if sheet.cell(row=row, column=col).value =="short" and canTradeShort == True:
			#print(row, col, "sell short")
			shortStockOnHand += ((currentCash/2)/sheet.cell(row=row, column =8).value)
			currentCash += (currentCash/2)
			#print(shortStockOnHand, currentCash)
			canTradeShort = False 

print("Conclusion 2017:")
#print("Starting Cash:", start17)
print("Current Cash:", currentCash)
if start17 > currentCash:
	print("Loss:", start17 - currentCash)
elif start17 < currentCash:
	print("Profit:", currentCash - start17)
else:
	print("Breakeven") 


#liquidate everything
fin = longStockOnHand* sheet.cell(row=sheet.max_row, column =8).value
currentCash += fin
longStockOnHand = 0
print('leftover stock profit:',fin)
print("=====Conclusion Total=====")

print("Starting Cash:", startingCash)
print("Current Cash:", currentCash)
if startingCash > currentCash:
	print("Loss:", startingCash - currentCash)
elif startingCash < currentCash:
	print("Profit:", currentCash - startingCash)
else:
	print("Breakeven")

